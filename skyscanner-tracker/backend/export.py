from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

EXPORT_DIR = Path(__file__).parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

COLUMNS = [
    ("Route", 22), ("Travel Date", 12), ("Airline", 18), ("Depart", 10),
    ("Arrive", 10), ("Duration", 10), ("Stops", 7), ("Stopover", 14),
    ("Price", 10), ("Currency", 9), ("Searched At", 18), ("Booking Link", 45),
]


def export_to_excel(rows: list[dict]) -> Path:
    """rows: list of dicts with keys matching the search+result join used by
    the /api/export endpoint. Writes a fresh timestamped .xlsx and returns
    its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Flights"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(rows, start=2):
        route = f"{r.get('origin_query')} -> {r.get('destination_query')}"
        values = [
            route, r.get("travel_date"), r.get("airline"), r.get("depart_time"),
            r.get("arrival_time"), r.get("duration"), r.get("stops"),
            r.get("stopover_airport"), r.get("price"), r.get("currency"),
            r.get("searched_at"),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        link = r.get("booking_link")
        link_cell = ws.cell(row=row_idx, column=len(COLUMNS), value=link or "")
        if link:
            link_cell.hyperlink = link
            link_cell.style = "Hyperlink"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = EXPORT_DIR / f"flights_{ts}.xlsx"
    wb.save(out_path)
    return out_path
